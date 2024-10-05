import os
import openpyxl

from utils import get_logger

logger = get_logger()

def get_or_create_excel_file(filename="record.xlsx"):
    if not os.path.exists(filename):
        # create a mew workbook if it does not exists
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Tracking"
        sheet.append(("User Profile", "Poll"))
        workbook.save(filename)

        logger.info(f"Excel file {filename} has been created successfully.")
    else:
        # load the existing workbook
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    
    return workbook, sheet

def update_excel_file(workbook, sheet, records, filename="record.xlsx"):
    for new_record in records:
        sheet.append(new_record)
    workbook.save(filename)

def get_visited_polls(sheet):
    return set(next(sheet.iter_cols(min_row=2, min_col=2, values_only=True)))

def get_visited_profiles(sheet):
    return set(next(sheet.iter_cols(min_row=2, min_col=1, values_only=True)))



